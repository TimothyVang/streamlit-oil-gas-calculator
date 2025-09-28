#!/usr/bin/env python3
"""
BAH Jackson Sands Oil & Gas Investment Analysis
Streamlit Web Application for Investor Presentations

Run with: streamlit run streamlit_oil_gas_app.py
"""

import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px
from datetime import datetime, timedelta
import base64
from io import BytesIO
import scipy.stats as stats
from scipy.optimize import fsolve

# Page configuration
st.set_page_config(
    page_title="BAH Jackson Sands - Oil & Gas Investment Analysis",
    page_icon="🛢️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for professional styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f4e79;
        text-align: center;
        margin-bottom: 2rem;
    }
    .metric-card {
        background: linear-gradient(90deg, #f0f2f6 0%, #ffffff 100%);
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #1f4e79;
        margin: 0.5rem 0;
    }
    .positive-metric {
        color: #28a745;
        font-weight: bold;
    }
    .negative-metric {
        color: #dc3545;
        font-weight: bold;
    }
    .sidebar .sidebar-content {
        background: linear-gradient(180deg, #1f4e79 0%, #2c5f82 100%);
    }
</style>
""", unsafe_allow_html=True)

def calculate_oil_gas_model(oil_price, gas_price, initial_production, decline_rate, discount_rate):
    """
    Calculate complete 60-month oil & gas financial model

    Returns: DataFrame with monthly calculations and summary metrics
    """

    # Convert percentages to decimals
    decline_rate = decline_rate / 100
    discount_rate = discount_rate / 100

    # Initialize results DataFrame
    months = list(range(1, 61))
    df = pd.DataFrame({'Month': months})

    # Calculate production decline
    df['Oil_Production'] = [initial_production * ((1 - decline_rate) ** (month - 1)) for month in months]
    df['Gas_Production'] = df['Oil_Production'] * 6  # 6:1 gas-to-oil ratio

    # Calculate revenues (in $000s)
    df['Oil_Revenue'] = df['Oil_Production'] * oil_price / 1000
    df['Gas_Revenue'] = df['Gas_Production'] * gas_price / 1000
    df['Total_Revenue'] = df['Oil_Revenue'] + df['Gas_Revenue']

    # Calculate operating expenses
    base_opex = 15  # $15k per month
    cost_escalation = 0.025  # 2.5% annual
    df['Operating_Expenses'] = [base_opex * ((1 + cost_escalation/12) ** (month - 1)) for month in months]
    df['Severance_Tax'] = df['Total_Revenue'] * 0.075  # 7.5% severance tax
    df['Total_OpEx'] = df['Operating_Expenses'] + df['Severance_Tax']

    # Calculate net operating income
    df['Net_Operating_Income'] = df['Total_Revenue'] - df['Total_OpEx']

    # Calculate CapEx schedule
    capex_schedule = [500, 300, 200, 100, 50, 50] + [10]*18 + [5]*36
    df['CapEx'] = capex_schedule

    # Calculate cash flows
    df['Net_Cash_Flow'] = df['Net_Operating_Income'] - df['CapEx']
    df['Cumulative_Cash_Flow'] = df['Net_Cash_Flow'].cumsum()

    # Calculate NPV
    monthly_discount = discount_rate / 12
    df['PV_Factor'] = [(1 + monthly_discount) ** -month for month in months]
    df['PV_Cash_Flow'] = df['Net_Cash_Flow'] * df['PV_Factor']

    # Summary metrics
    total_investment = df['CapEx'].sum()
    total_revenue = df['Total_Revenue'].sum()
    npv = df['PV_Cash_Flow'].sum()
    final_cumulative = df['Cumulative_Cash_Flow'].iloc[-1]

    # Calculate payback period
    payback_month = 60  # Default if never pays back
    for i, cum_cf in enumerate(df['Cumulative_Cash_Flow']):
        if cum_cf >= 0:
            payback_month = i + 1
            break

    # Calculate IRR using improved method
    cash_flows = df['Net_Cash_Flow'].values
    irr = calculate_irr_scipy(cash_flows)

    summary = {
        'Total_Investment': total_investment,
        'Total_Revenue': total_revenue,
        'NPV': npv,
        'IRR': irr,
        'Payback_Months': payback_month,
        'Final_Cumulative_CF': final_cumulative,
        'Peak_Production': df['Oil_Production'].max(),
        'Final_Production': df['Oil_Production'].iloc[-1]
    }

    return df, summary

def create_production_chart(df):
    """Create production decline chart"""
    fig = go.Figure()

    # Oil production
    fig.add_trace(go.Scatter(
        x=df['Month'],
        y=df['Oil_Production'],
        mode='lines',
        name='Oil Production (bbl/month)',
        line=dict(color='#1f77b4', width=3),
        hovertemplate='Month %{x}<br>Oil: %{y:,.0f} bbl/month<extra></extra>'
    ))

    # Gas production (secondary axis)
    fig.add_trace(go.Scatter(
        x=df['Month'],
        y=df['Gas_Production'],
        mode='lines',
        name='Gas Production (MCF/month)',
        line=dict(color='#ff7f0e', width=3),
        yaxis='y2',
        hovertemplate='Month %{x}<br>Gas: %{y:,.0f} MCF/month<extra></extra>'
    ))

    fig.update_layout(
        title='Production Decline Curves',
        xaxis_title='Month',
        yaxis_title='Oil Production (bbl/month)',
        yaxis2=dict(
            title='Gas Production (MCF/month)',
            overlaying='y',
            side='right'
        ),
        hovermode='x unified',
        height=400
    )

    return fig

def create_revenue_chart(df):
    """Create revenue stream chart"""
    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=df['Month'],
        y=df['Oil_Revenue'],
        mode='lines',
        name='Oil Revenue',
        line=dict(color='#2ca02c', width=3),
        stackgroup='one',
        hovertemplate='Month %{x}<br>Oil Revenue: $%{y:,.0f}k<extra></extra>'
    ))

    fig.add_trace(go.Scatter(
        x=df['Month'],
        y=df['Gas_Revenue'],
        mode='lines',
        name='Gas Revenue',
        line=dict(color='#d62728', width=3),
        stackgroup='one',
        hovertemplate='Month %{x}<br>Gas Revenue: $%{y:,.0f}k<extra></extra>'
    ))

    fig.update_layout(
        title='Revenue Streams Over Time',
        xaxis_title='Month',
        yaxis_title='Revenue ($000s)',
        hovermode='x unified',
        height=400
    )

    return fig

def create_cashflow_chart(df):
    """Create cash flow analysis chart"""
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=('Monthly Net Cash Flow', 'Cumulative Cash Flow'),
        vertical_spacing=0.1
    )

    # Monthly cash flow
    colors = ['red' if x < 0 else 'green' for x in df['Net_Cash_Flow']]
    fig.add_trace(
        go.Bar(
            x=df['Month'],
            y=df['Net_Cash_Flow'],
            marker_color=colors,
            name='Monthly Net Cash Flow',
            hovertemplate='Month %{x}<br>Net CF: $%{y:,.0f}k<extra></extra>'
        ),
        row=1, col=1
    )

    # Cumulative cash flow
    fig.add_trace(
        go.Scatter(
            x=df['Month'],
            y=df['Cumulative_Cash_Flow'],
            mode='lines',
            name='Cumulative Cash Flow',
            line=dict(color='#9467bd', width=3),
            fill='tozeroy',
            hovertemplate='Month %{x}<br>Cumulative CF: $%{y:,.0f}k<extra></extra>'
        ),
        row=2, col=1
    )

    # Add break-even line
    fig.add_hline(y=0, line_dash="dash", line_color="black", row=2, col=1)

    fig.update_layout(
        height=600,
        showlegend=False
    )

    fig.update_xaxes(title_text="Month", row=2, col=1)
    fig.update_yaxes(title_text="Cash Flow ($000s)", row=1, col=1)
    fig.update_yaxes(title_text="Cumulative CF ($000s)", row=2, col=1)

    return fig

def get_investment_grade(npv_return):
    """Determine investment grade based on NPV return"""
    if npv_return > 25:
        return "🏆 EXCELLENT", "#28a745"
    elif npv_return > 15:
        return "🥇 VERY GOOD", "#28a745"
    elif npv_return > 5:
        return "🥈 GOOD", "#ffc107"
    elif npv_return > 0:
        return "🥉 ACCEPTABLE", "#ffc107"
    else:
        return "❌ POOR", "#dc3545"

def calculate_irr_scipy(cash_flows):
    """Calculate IRR using scipy optimization"""
    def npv_func(rate):
        return sum(cf / (1 + rate/12)**(i+1) for i, cf in enumerate(cash_flows))

    try:
        irr = fsolve(npv_func, 0.1)[0] * 100  # Convert to percentage
        return max(min(irr, 1000), -100)  # Cap at reasonable bounds
    except:
        return 0

def run_monte_carlo_simulation(base_params, num_simulations=1000, volatility_factors=None):
    """
    Run Monte Carlo simulation with stochastic variables

    Parameters:
    - base_params: dict with oil_price, gas_price, initial_production, decline_rate, discount_rate
    - num_simulations: number of Monte Carlo runs
    - volatility_factors: dict with standard deviations for each parameter
    """

    if volatility_factors is None:
        volatility_factors = {
            'oil_price': 0.15,      # 15% volatility
            'gas_price': 0.25,      # 25% volatility
            'initial_production': 0.10,  # 10% volatility
            'decline_rate': 0.20,   # 20% volatility
        }

    results = []

    # Progress bar for Monte Carlo
    progress_bar = st.progress(0)
    status_text = st.empty()

    for i in range(num_simulations):
        # Generate random parameters based on normal distributions
        oil_price = max(20, np.random.normal(base_params['oil_price'],
                                           base_params['oil_price'] * volatility_factors['oil_price']))
        gas_price = max(1, np.random.normal(base_params['gas_price'],
                                          base_params['gas_price'] * volatility_factors['gas_price']))
        initial_production = max(100, np.random.normal(base_params['initial_production'],
                                                     base_params['initial_production'] * volatility_factors['initial_production']))
        decline_rate = max(0.5, min(10, np.random.normal(base_params['decline_rate'],
                                                        base_params['decline_rate'] * volatility_factors['decline_rate'])))

        # Run calculation with random parameters
        _, summary = calculate_oil_gas_model(oil_price, gas_price, initial_production,
                                           decline_rate, base_params['discount_rate'])

        # Store results
        results.append({
            'Oil_Price': oil_price,
            'Gas_Price': gas_price,
            'Initial_Production': initial_production,
            'Decline_Rate': decline_rate,
            'NPV': summary['NPV'],
            'IRR': summary['IRR'],
            'Payback_Months': summary['Payback_Months'],
            'Final_Cumulative_CF': summary['Final_Cumulative_CF'],
            'Total_Investment': summary['Total_Investment'],
            'Total_Revenue': summary['Total_Revenue']
        })

        # Update progress
        progress = (i + 1) / num_simulations
        progress_bar.progress(progress)
        status_text.text(f'Running Monte Carlo simulation: {i+1}/{num_simulations}')

    progress_bar.empty()
    status_text.empty()

    return pd.DataFrame(results)

def create_monte_carlo_charts(mc_results):
    """Create Monte Carlo analysis charts"""

    # NPV distribution histogram
    fig_npv = go.Figure()
    fig_npv.add_trace(go.Histogram(
        x=mc_results['NPV'],
        nbinsx=50,
        name='NPV Distribution',
        marker_color='rgba(55, 128, 191, 0.7)',
        hovertemplate='NPV Range: %{x}<br>Count: %{y}<extra></extra>'
    ))

    # Add percentile lines
    npv_p10 = np.percentile(mc_results['NPV'], 10)
    npv_p50 = np.percentile(mc_results['NPV'], 50)
    npv_p90 = np.percentile(mc_results['NPV'], 90)

    fig_npv.add_vline(x=npv_p10, line_dash="dash", line_color="red",
                      annotation_text=f"P10: ${npv_p10:,.0f}k")
    fig_npv.add_vline(x=npv_p50, line_dash="dash", line_color="green",
                      annotation_text=f"P50: ${npv_p50:,.0f}k")
    fig_npv.add_vline(x=npv_p90, line_dash="dash", line_color="blue",
                      annotation_text=f"P90: ${npv_p90:,.0f}k")

    fig_npv.update_layout(
        title='NPV Distribution (Monte Carlo)',
        xaxis_title='NPV ($000s)',
        yaxis_title='Frequency',
        height=400
    )

    # Correlation matrix
    correlation_vars = ['Oil_Price', 'Gas_Price', 'Initial_Production', 'Decline_Rate', 'NPV', 'IRR']
    correlation_matrix = mc_results[correlation_vars].corr()

    fig_corr = go.Figure(data=go.Heatmap(
        z=correlation_matrix.values,
        x=correlation_matrix.columns,
        y=correlation_matrix.columns,
        colorscale='RdBu',
        zmid=0,
        text=correlation_matrix.round(2).values,
        texttemplate="%{text}",
        textfont={"size": 10},
        hovertemplate='%{x} vs %{y}<br>Correlation: %{z:.2f}<extra></extra>'
    ))

    fig_corr.update_layout(
        title='Parameter Correlation Matrix',
        height=400
    )

    # Sensitivity tornado chart
    # Calculate correlation of each input with NPV
    correlations = {
        'Oil Price': mc_results['Oil_Price'].corr(mc_results['NPV']),
        'Gas Price': mc_results['Gas_Price'].corr(mc_results['NPV']),
        'Initial Production': mc_results['Initial_Production'].corr(mc_results['NPV']),
        'Decline Rate': mc_results['Decline_Rate'].corr(mc_results['NPV'])
    }

    # Sort by absolute correlation
    sorted_corr = sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True)

    fig_tornado = go.Figure()

    for i, (param, corr) in enumerate(sorted_corr):
        color = 'green' if corr > 0 else 'red'
        fig_tornado.add_trace(go.Bar(
            y=[param],
            x=[abs(corr)],
            orientation='h',
            name=param,
            marker_color=color,
            text=f'{corr:.3f}',
            textposition='inside',
            hovertemplate=f'{param}<br>Correlation with NPV: {corr:.3f}<extra></extra>'
        ))

    fig_tornado.update_layout(
        title='Sensitivity Analysis (Correlation with NPV)',
        xaxis_title='Absolute Correlation with NPV',
        yaxis_title='Parameters',
        height=300,
        showlegend=False
    )

    return fig_npv, fig_corr, fig_tornado

def create_risk_assessment(mc_results):
    """Create risk assessment metrics"""
    npv_values = mc_results['NPV']

    # Calculate risk metrics
    probability_positive = (npv_values > 0).mean() * 100
    probability_excellent = (npv_values > npv_values.quantile(0.75)).mean() * 100
    value_at_risk_5 = np.percentile(npv_values, 5)
    expected_value = npv_values.mean()

    # Standard deviation and coefficient of variation
    std_dev = npv_values.std()
    cv = std_dev / abs(expected_value) if expected_value != 0 else float('inf')

    # Downside deviation (volatility of negative returns)
    downside_returns = npv_values[npv_values < expected_value]
    downside_deviation = downside_returns.std() if len(downside_returns) > 0 else 0

    # Sharpe-like ratio (excess return per unit of risk)
    sharpe_ratio = expected_value / std_dev if std_dev > 0 else 0

    return {
        'Expected_NPV': expected_value,
        'Probability_Positive': probability_positive,
        'Probability_Excellent': probability_excellent,
        'Value_at_Risk_5': value_at_risk_5,
        'Standard_Deviation': std_dev,
        'Coefficient_of_Variation': cv,
        'Downside_Deviation': downside_deviation,
        'Risk_Adjusted_Return': sharpe_ratio,
        'P10_NPV': np.percentile(npv_values, 10),
        'P50_NPV': np.percentile(npv_values, 50),
        'P90_NPV': np.percentile(npv_values, 90)
    }

def export_to_excel(df, summary, params, mc_results=None, risk_metrics=None):
    """Export results to professionally formatted Excel file"""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.formatting.rule import ColorScaleRule

        # Define professional styling
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        data_font = Font(name='Calibri', size=11)
        currency_font = Font(name='Calibri', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # 1. Executive Summary Sheet
        exec_summary = writer.book.create_sheet('Executive_Summary', 0)
        exec_summary['A1'] = 'BAH Jackson Sands - Oil & Gas Investment Analysis'
        exec_summary['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
        exec_summary['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
        exec_summary['A2'].font = Font(name='Calibri', size=10, italic=True)

        # Key metrics in executive summary
        exec_data = [
            ['KEY INVESTMENT METRICS', ''],
            ['Total Investment Required', f"${summary['Total_Investment']:,.0f}k"],
            ['Net Present Value (NPV)', f"${summary['NPV']:,.0f}k"],
            ['Internal Rate of Return (IRR)', f"{summary['IRR']:.1f}%"],
            ['Payback Period', f"{summary['Payback_Months']:.0f} months"],
            ['Final Cumulative Cash Flow', f"${summary['Final_Cumulative_CF']:,.0f}k"],
            ['', ''],
            ['PRODUCTION SUMMARY', ''],
            ['Peak Oil Production', f"{summary['Peak_Production']:,.0f} bbl/month"],
            ['Final Oil Production', f"{summary['Final_Production']:,.0f} bbl/month"],
            ['Total Revenue (60 months)', f"${summary['Total_Revenue']:,.0f}k"]
        ]

        for row_idx, (metric, value) in enumerate(exec_data, start=4):
            exec_summary[f'A{row_idx}'] = metric
            exec_summary[f'B{row_idx}'] = value
            if metric.isupper():  # Section headers
                exec_summary[f'A{row_idx}'].font = Font(name='Calibri', size=12, bold=True, color='1F4E79')
            exec_summary[f'A{row_idx}'].font = Font(name='Calibri', size=11, bold=metric != '')

        # Auto-size columns
        exec_summary.column_dimensions['A'].width = 30
        exec_summary.column_dimensions['B'].width = 20

        # 2. Monthly Analysis Sheet with formatting
        monthly_df = df.copy()
        # Format columns for better readability
        monthly_df = monthly_df.round({
            'Oil_Production': 0, 'Gas_Production': 0,
            'Oil_Revenue': 1, 'Gas_Revenue': 1, 'Total_Revenue': 1,
            'Operating_Expenses': 1, 'Severance_Tax': 1, 'Total_OpEx': 1,
            'Net_Operating_Income': 1, 'CapEx': 1, 'Net_Cash_Flow': 1,
            'Cumulative_Cash_Flow': 1, 'PV_Cash_Flow': 1
        })

        # Rename columns for professional presentation
        monthly_df = monthly_df.rename(columns={
            'Oil_Production': 'Oil Production (bbl)',
            'Gas_Production': 'Gas Production (MCF)',
            'Oil_Revenue': 'Oil Revenue ($k)',
            'Gas_Revenue': 'Gas Revenue ($k)',
            'Total_Revenue': 'Total Revenue ($k)',
            'Operating_Expenses': 'Operating Expenses ($k)',
            'Severance_Tax': 'Severance Tax ($k)',
            'Total_OpEx': 'Total OpEx ($k)',
            'Net_Operating_Income': 'Net Operating Income ($k)',
            'CapEx': 'Capital Expenditure ($k)',
            'Net_Cash_Flow': 'Net Cash Flow ($k)',
            'Cumulative_Cash_Flow': 'Cumulative Cash Flow ($k)',
            'PV_Factor': 'PV Factor',
            'PV_Cash_Flow': 'PV Cash Flow ($k)'
        })

        monthly_df.to_excel(writer, sheet_name='Monthly_Analysis', index=False, startrow=1)
        monthly_ws = writer.sheets['Monthly_Analysis']

        # Add title and format headers
        monthly_ws['A1'] = 'Monthly Financial Analysis - 60 Month Projection'
        monthly_ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')

        # Format headers
        for col in range(1, len(monthly_df.columns) + 1):
            cell = monthly_ws.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Format data cells
        for row in range(3, len(monthly_df) + 3):
            for col in range(1, len(monthly_df.columns) + 1):
                cell = monthly_ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = border
                # Format currency columns
                if 'Revenue' in monthly_df.columns[col-1] or 'Cash Flow' in monthly_df.columns[col-1] or 'Expense' in monthly_df.columns[col-1] or 'CapEx' in monthly_df.columns[col-1]:
                    cell.number_format = '#,##0.0'

        # Auto-size columns
        for column in monthly_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            monthly_ws.column_dimensions[column_letter].width = adjusted_width

        # 3. Investment Summary Sheet
        summary_data = []
        for key, value in summary.items():
            formatted_key = key.replace('_', ' ').title()
            if isinstance(value, (int, float)):
                if 'NPV' in key or 'Revenue' in key or 'Investment' in key or 'CF' in key:
                    formatted_value = f"${value:,.0f}k"
                elif 'IRR' in key:
                    formatted_value = f"{value:.1f}%"
                elif 'Months' in key:
                    formatted_value = f"{value:.0f} months"
                elif 'Production' in key:
                    formatted_value = f"{value:,.0f} bbl/month"
                else:
                    formatted_value = f"{value:,.2f}"
            else:
                formatted_value = str(value)
            summary_data.append([formatted_key, formatted_value])

        summary_df = pd.DataFrame(summary_data, columns=['Investment Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Investment_Summary', index=False, startrow=1)
        summary_ws = writer.sheets['Investment_Summary']

        # Format Investment Summary sheet
        summary_ws['A1'] = 'Investment Performance Summary'
        summary_ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')

        # Format headers and data
        for col in range(1, 3):
            cell = summary_ws.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        for row in range(3, len(summary_df) + 3):
            for col in range(1, 3):
                cell = summary_ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = border

        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 20

        # 4. Parameters Sheet
        params_data = []
        for key, value in params.items():
            if isinstance(value, (int, float)):
                if 'Price' in key and '$' not in key:
                    formatted_value = f"${value:.2f}"
                elif '%' in key:
                    formatted_value = f"{value:.1f}%"
                else:
                    formatted_value = f"{value:,.0f}"
            else:
                formatted_value = str(value)
            params_data.append([key, formatted_value])

        params_df = pd.DataFrame(params_data, columns=['Parameter', 'Value'])
        params_df.to_excel(writer, sheet_name='Input_Parameters', index=False, startrow=1)
        params_ws = writer.sheets['Input_Parameters']

        # Format Parameters sheet
        params_ws['A1'] = 'Model Input Parameters'
        params_ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')

        # Format headers and data
        for col in range(1, 3):
            cell = params_ws.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        for row in range(3, len(params_df) + 3):
            for col in range(1, 3):
                cell = params_ws.cell(row=row, column=col)
                cell.font = data_font
                cell.border = border

        params_ws.column_dimensions['A'].width = 35
        params_ws.column_dimensions['B'].width = 20

        # 5. Monte Carlo Results (if available)
        if mc_results is not None:
            # Sample of Monte Carlo results for Excel (first 1000 rows to avoid file size issues)
            mc_sample = mc_results.head(1000).round(2)
            mc_sample.to_excel(writer, sheet_name='Monte_Carlo_Sample', index=False, startrow=1)
            mc_ws = writer.sheets['Monte_Carlo_Sample']

            mc_ws['A1'] = f'Monte Carlo Simulation Results (Sample of {len(mc_sample)} runs)'
            mc_ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')

            # Format headers
            for col in range(1, len(mc_sample.columns) + 1):
                cell = mc_ws.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

        # 6. Risk Assessment (if available)
        if risk_metrics is not None:
            risk_data = []
            for key, value in risk_metrics.items():
                formatted_key = key.replace('_', ' ').title()
                if isinstance(value, (int, float)):
                    if 'NPV' in key:
                        formatted_value = f"${value:,.0f}k"
                    elif 'Probability' in key or 'Coefficient' in key:
                        formatted_value = f"{value:.1f}%"
                    elif 'Return' in key or 'Deviation' in key:
                        formatted_value = f"{value:.2f}"
                    else:
                        formatted_value = f"{value:,.2f}"
                else:
                    formatted_value = str(value)
                risk_data.append([formatted_key, formatted_value])

            risk_df = pd.DataFrame(risk_data, columns=['Risk Metric', 'Value'])
            risk_df.to_excel(writer, sheet_name='Risk_Assessment', index=False, startrow=1)
            risk_ws = writer.sheets['Risk_Assessment']

            risk_ws['A1'] = 'Monte Carlo Risk Assessment'
            risk_ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1F4E79')

            # Format headers and data
            for col in range(1, 3):
                cell = risk_ws.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border

            for row in range(3, len(risk_df) + 3):
                for col in range(1, 3):
                    cell = risk_ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border

            risk_ws.column_dimensions['A'].width = 30
            risk_ws.column_dimensions['B'].width = 20

    output.seek(0)
    return output

def export_to_csv_professional(df, summary, params, mc_results=None, risk_metrics=None):
    """Export results to professional CSV with multiple sections"""

    # Create a comprehensive dataset
    csv_content = []

    # Header section
    csv_content.append(['BAH Jackson Sands - Oil & Gas Investment Analysis'])
    csv_content.append([f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'])
    csv_content.append([''])

    # Executive Summary
    csv_content.append(['=== EXECUTIVE SUMMARY ==='])
    csv_content.append(['Metric', 'Value'])
    csv_content.append(['Total Investment Required', f"${summary['Total_Investment']:,.0f}k"])
    csv_content.append(['Net Present Value (NPV)', f"${summary['NPV']:,.0f}k"])
    csv_content.append(['Internal Rate of Return (IRR)', f"{summary['IRR']:.1f}%"])
    csv_content.append(['Payback Period', f"{summary['Payback_Months']:.0f} months"])
    csv_content.append(['Final Cumulative Cash Flow', f"${summary['Final_Cumulative_CF']:,.0f}k"])
    csv_content.append(['Peak Oil Production', f"{summary['Peak_Production']:,.0f} bbl/month"])
    csv_content.append(['Final Oil Production', f"{summary['Final_Production']:,.0f} bbl/month"])
    csv_content.append(['Total Revenue (60 months)', f"${summary['Total_Revenue']:,.0f}k"])
    csv_content.append([''])

    # Input Parameters
    csv_content.append(['=== INPUT PARAMETERS ==='])
    csv_content.append(['Parameter', 'Value'])
    for key, value in params.items():
        if isinstance(value, (int, float)):
            if 'Price' in key and '$' not in key:
                formatted_value = f"${value:.2f}"
            elif '%' in key:
                formatted_value = f"{value:.1f}%"
            else:
                formatted_value = f"{value:,.0f}"
        else:
            formatted_value = str(value)
        csv_content.append([key, formatted_value])
    csv_content.append([''])

    # Monthly Analysis
    csv_content.append(['=== MONTHLY FINANCIAL ANALYSIS ==='])

    # Format the monthly data
    monthly_df = df.copy()
    monthly_df = monthly_df.round({
        'Oil_Production': 0, 'Gas_Production': 0,
        'Oil_Revenue': 1, 'Gas_Revenue': 1, 'Total_Revenue': 1,
        'Operating_Expenses': 1, 'Severance_Tax': 1, 'Total_OpEx': 1,
        'Net_Operating_Income': 1, 'CapEx': 1, 'Net_Cash_Flow': 1,
        'Cumulative_Cash_Flow': 1, 'PV_Cash_Flow': 1
    })

    # Rename columns for professional presentation
    monthly_df = monthly_df.rename(columns={
        'Oil_Production': 'Oil Production (bbl)',
        'Gas_Production': 'Gas Production (MCF)',
        'Oil_Revenue': 'Oil Revenue ($k)',
        'Gas_Revenue': 'Gas Revenue ($k)',
        'Total_Revenue': 'Total Revenue ($k)',
        'Operating_Expenses': 'Operating Expenses ($k)',
        'Severance_Tax': 'Severance Tax ($k)',
        'Total_OpEx': 'Total OpEx ($k)',
        'Net_Operating_Income': 'Net Operating Income ($k)',
        'CapEx': 'Capital Expenditure ($k)',
        'Net_Cash_Flow': 'Net Cash Flow ($k)',
        'Cumulative_Cash_Flow': 'Cumulative Cash Flow ($k)',
        'PV_Factor': 'PV Factor',
        'PV_Cash_Flow': 'PV Cash Flow ($k)'
    })

    # Add headers
    csv_content.append(list(monthly_df.columns))

    # Add data rows
    for _, row in monthly_df.iterrows():
        csv_content.append(list(row))

    # Add risk metrics if available
    if risk_metrics is not None:
        csv_content.append([''])
        csv_content.append(['=== RISK ASSESSMENT METRICS ==='])
        csv_content.append(['Risk Metric', 'Value'])
        for key, value in risk_metrics.items():
            formatted_key = key.replace('_', ' ').title()
            if isinstance(value, (int, float)):
                if 'NPV' in key:
                    formatted_value = f"${value:,.0f}k"
                elif 'Probability' in key:
                    formatted_value = f"{value:.1f}%"
                elif 'Return' in key or 'Deviation' in key:
                    formatted_value = f"{value:.2f}"
                else:
                    formatted_value = f"{value:,.2f}"
            else:
                formatted_value = str(value)
            csv_content.append([formatted_key, formatted_value])

    # Convert to CSV string
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)
    for row in csv_content:
        writer.writerow(row)

    return output.getvalue()

# Main Application
def main():
    # Header
    st.markdown('<h1 class="main-header">🛢️ BAH Jackson Sands Oil & Gas Investment Analysis</h1>', unsafe_allow_html=True)
    st.markdown("### Interactive Financial Model for Investment Decision Making")

    # Sidebar for parameters
    st.sidebar.header("🎛️ Investment Parameters")
    st.sidebar.markdown("Adjust the parameters below to analyze different scenarios:")

    # Market parameters
    st.sidebar.subheader("💰 Market Conditions")
    oil_price = st.sidebar.slider("Oil Price ($/barrel)", 30.0, 120.0, 65.0, 1.0)
    gas_price = st.sidebar.slider("Gas Price ($/MCF)", 1.0, 8.0, 3.25, 0.25)

    # Production parameters
    st.sidebar.subheader("📉 Production Profile")
    initial_production = st.sidebar.slider("Initial Production (bbl/month)", 100, 3000, 1000, 50)
    decline_rate = st.sidebar.slider("Monthly Decline Rate (%)", 0.5, 5.0, 1.5, 0.1)

    # Financial parameters
    st.sidebar.subheader("📊 Financial Assumptions")
    discount_rate = st.sidebar.slider("Discount Rate (%)", 5.0, 20.0, 10.0, 0.5)

    # Calculate model
    df, summary = calculate_oil_gas_model(oil_price, gas_price, initial_production, decline_rate, discount_rate)

    # Key metrics display
    st.header("📊 Key Investment Metrics")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(
            label="💰 Total Investment",
            value=f"${summary['Total_Investment']:,.0f}k"
        )

    with col2:
        st.metric(
            label="🎯 NPV @ " + f"{discount_rate:.1f}%",
            value=f"${summary['NPV']:,.0f}k",
            delta=f"{(summary['NPV']/summary['Total_Investment']*100):.1f}% return"
        )

    with col3:
        st.metric(
            label="⏱️ Payback Period",
            value=f"{summary['Payback_Months']:.0f} months",
            delta=f"{summary['Payback_Months']/12:.1f} years"
        )

    with col4:
        st.metric(
            label="💵 Final Cumulative CF",
            value=f"${summary['Final_Cumulative_CF']:,.0f}k"
        )

    # Investment grade assessment
    npv_return = (summary['NPV'] / summary['Total_Investment']) * 100
    grade, color = get_investment_grade(npv_return)

    st.markdown(f"""
    <div style="background: linear-gradient(90deg, {color}20 0%, {color}10 100%);
                padding: 1rem; border-radius: 10px; border-left: 5px solid {color}; margin: 1rem 0;">
        <h3 style="color: {color}; margin: 0;">Investment Grade: {grade}</h3>
        <p style="margin: 0.5rem 0 0 0;">NPV Return: {npv_return:.1f}% | IRR: {summary['IRR']:.1f}%</p>
    </div>
    """, unsafe_allow_html=True)

    # Charts section
    st.header("📈 Financial Analysis Charts")

    # Production analysis
    st.subheader("Production Decline Analysis")
    fig_production = create_production_chart(df)
    st.plotly_chart(fig_production, use_container_width=True)

    col1, col2 = st.columns(2)

    with col1:
        st.info(f"**Peak Production:** {summary['Peak_Production']:,.0f} bbl/month")
        st.info(f"**Final Production:** {summary['Final_Production']:,.0f} bbl/month")

    with col2:
        decline_total = (1 - summary['Final_Production']/summary['Peak_Production']) * 100
        st.info(f"**Total Decline:** {decline_total:.1f}% over 60 months")
        st.info(f"**Average Daily Production:** {summary['Peak_Production']/30:.0f} bbl/day initial")

    # Revenue analysis
    st.subheader("Revenue Stream Analysis")
    fig_revenue = create_revenue_chart(df)
    st.plotly_chart(fig_revenue, use_container_width=True)

    col1, col2 = st.columns(2)
    with col1:
        st.info(f"**Total Revenue (60 months):** ${summary['Total_Revenue']:,.0f}k")
        oil_percentage = (df['Oil_Revenue'].sum() / summary['Total_Revenue']) * 100
        st.info(f"**Oil Revenue Share:** {oil_percentage:.1f}%")

    with col2:
        gas_percentage = (df['Gas_Revenue'].sum() / summary['Total_Revenue']) * 100
        st.info(f"**Gas Revenue Share:** {gas_percentage:.1f}%")
        st.info(f"**Average Monthly Revenue:** ${summary['Total_Revenue']/60:,.0f}k")

    # Cash flow analysis
    st.subheader("Cash Flow Analysis")
    fig_cashflow = create_cashflow_chart(df)
    st.plotly_chart(fig_cashflow, use_container_width=True)

    # Detailed data table
    st.header("📋 Detailed Monthly Analysis")

    # Show/hide detailed table
    if st.checkbox("Show Detailed Monthly Breakdown"):
        # Format dataframe for display
        display_df = df.copy()

        # Round numerical columns
        numeric_columns = ['Oil_Production', 'Gas_Production', 'Oil_Revenue', 'Gas_Revenue',
                          'Total_Revenue', 'Total_OpEx', 'Net_Operating_Income', 'CapEx',
                          'Net_Cash_Flow', 'Cumulative_Cash_Flow']

        for col in numeric_columns:
            if col in ['Oil_Production', 'Gas_Production']:
                display_df[col] = display_df[col].round(0).astype(int)
            else:
                display_df[col] = display_df[col].round(1)

        # Select key columns for display
        key_columns = ['Month', 'Oil_Production', 'Total_Revenue', 'Total_OpEx',
                      'Net_Operating_Income', 'CapEx', 'Net_Cash_Flow', 'Cumulative_Cash_Flow']

        # Rename for better display
        display_df = display_df.rename(columns={
            'Oil_Production': 'Oil Prod (bbl)',
            'Total_Revenue': 'Revenue ($k)',
            'Total_OpEx': 'OpEx ($k)',
            'Net_Operating_Income': 'Net Income ($k)',
            'CapEx': 'CapEx ($k)',
            'Net_Cash_Flow': 'Net CF ($k)',
            'Cumulative_Cash_Flow': 'Cumulative CF ($k)'
        })

        st.dataframe(display_df[['Month', 'Oil Prod (bbl)', 'Revenue ($k)', 'OpEx ($k)',
                                'Net Income ($k)', 'CapEx ($k)', 'Net CF ($k)', 'Cumulative CF ($k)']],
                    use_container_width=True)

    # Scenario comparison
    st.header("🎯 Scenario Comparison")

    if st.button("🎯 Run Scenario Analysis"):
        scenarios = {
            'Conservative': {'oil_price': oil_price*0.8, 'gas_price': gas_price*0.8,
                           'initial_production': initial_production*0.8, 'decline_rate': decline_rate*1.5, 'discount_rate': discount_rate},
            'Base Case': {'oil_price': oil_price, 'gas_price': gas_price,
                         'initial_production': initial_production, 'decline_rate': decline_rate, 'discount_rate': discount_rate},
            'Optimistic': {'oil_price': oil_price*1.2, 'gas_price': gas_price*1.2,
                          'initial_production': initial_production*1.2, 'decline_rate': decline_rate*0.7, 'discount_rate': discount_rate}
        }

        scenario_results = {}
        for name, params in scenarios.items():
            _, summary_scenario = calculate_oil_gas_model(**params)
            scenario_results[name] = summary_scenario

        # Create scenario comparison chart
        scenario_names = list(scenario_results.keys())
        npv_values = [scenario_results[name]['NPV'] for name in scenario_names]

        fig_scenarios = go.Figure(data=[
            go.Bar(
                x=scenario_names,
                y=npv_values,
                marker_color=['#ff7f7f' if x < 0 else '#7fbf7f' for x in npv_values],
                text=[f'${x:,.0f}k' for x in npv_values],
                textposition='auto'
            )
        ])

        fig_scenarios.update_layout(
            title='NPV Comparison Across Scenarios',
            xaxis_title='Scenario',
            yaxis_title='NPV ($000s)',
            height=400
        )

        st.plotly_chart(fig_scenarios, use_container_width=True)

        # Scenario comparison table
        comparison_df = pd.DataFrame(scenario_results).T
        comparison_df = comparison_df[['Total_Investment', 'NPV', 'Payback_Months', 'Final_Cumulative_CF']].round(0)
        comparison_df.columns = ['Investment ($k)', 'NPV ($k)', 'Payback (months)', 'Final CF ($k)']

        st.subheader("Scenario Comparison Table")
        st.dataframe(comparison_df, use_container_width=True)

    # Monte Carlo Risk Analysis
    st.header("🎲 Monte Carlo Risk Analysis")
    st.markdown("**Advanced probabilistic analysis with thousands of simulations to quantify investment risk and uncertainty**")

    col1, col2 = st.columns(2)

    with col1:
        num_simulations = st.selectbox("Number of Simulations", [500, 1000, 2000, 5000], index=1)

    with col2:
        if st.button("🎲 Run Monte Carlo Analysis"):
            st.info("⏳ Running probabilistic analysis... This may take a moment.")

            # Prepare base parameters
            base_params = {
                'oil_price': oil_price,
                'gas_price': gas_price,
                'initial_production': initial_production,
                'decline_rate': decline_rate,
                'discount_rate': discount_rate
            }

            # Volatility assumptions section
            volatility_factors = None
            with st.expander("🔧 Volatility Assumptions", expanded=False):
                st.markdown("**Adjust the volatility (standard deviation) for each parameter:**")

                col1_vol, col2_vol = st.columns(2)

                with col1_vol:
                    oil_vol = st.slider("Oil Price Volatility (%)", 5, 50, 15, 1) / 100
                    gas_vol = st.slider("Gas Price Volatility (%)", 10, 60, 25, 1) / 100

                with col2_vol:
                    prod_vol = st.slider("Production Volatility (%)", 5, 30, 10, 1) / 100
                    decline_vol = st.slider("Decline Rate Volatility (%)", 10, 50, 20, 1) / 100

                volatility_factors = {
                    'oil_price': oil_vol,
                    'gas_price': gas_vol,
                    'initial_production': prod_vol,
                    'decline_rate': decline_vol
                }

            # Use default volatility factors if not set
            if volatility_factors is None:
                volatility_factors = {
                    'oil_price': 0.15,
                    'gas_price': 0.25,
                    'initial_production': 0.10,
                    'decline_rate': 0.20
                }

            # Run Monte Carlo simulation
            mc_results = run_monte_carlo_simulation(base_params, num_simulations, volatility_factors)

            # Calculate risk metrics
            risk_metrics = create_risk_assessment(mc_results)

            # Display risk assessment
            st.subheader("📊 Risk Assessment Summary")

            col1, col2, col3, col4 = st.columns(4)

            with col1:
                st.metric(
                    label="🎯 Expected NPV",
                    value=f"${risk_metrics['Expected_NPV']:,.0f}k",
                    delta=f"{risk_metrics['Probability_Positive']:.1f}% positive"
                )

            with col2:
                st.metric(
                    label="📉 Value at Risk (5%)",
                    value=f"${risk_metrics['Value_at_Risk_5']:,.0f}k",
                    delta=f"5% chance of worse outcome"
                )

            with col3:
                st.metric(
                    label="📈 Upside Potential (P90)",
                    value=f"${risk_metrics['P90_NPV']:,.0f}k",
                    delta=f"10% chance of better outcome"
                )

            with col4:
                st.metric(
                    label="⚖️ Risk-Adjusted Return",
                    value=f"{risk_metrics['Risk_Adjusted_Return']:.2f}",
                    delta=f"CV: {risk_metrics['Coefficient_of_Variation']:.2f}"
                )

            # Risk Assessment Cards
            st.subheader("🎯 Probability Analysis")

            col1, col2, col3 = st.columns(3)

            with col1:
                prob_positive = risk_metrics['Probability_Positive']
                color = "#28a745" if prob_positive > 70 else "#ffc107" if prob_positive > 50 else "#dc3545"
                st.markdown(f"""
                <div style="background: linear-gradient(90deg, {color}20 0%, {color}10 100%);
                            padding: 1rem; border-radius: 10px; border-left: 5px solid {color};">
                    <h4 style="color: {color}; margin: 0;">Probability of Positive NPV</h4>
                    <h2 style="color: {color}; margin: 0.5rem 0;">{prob_positive:.1f}%</h2>
                    <p style="margin: 0;">Chance of profitable investment</p>
                </div>
                """, unsafe_allow_html=True)

            with col2:
                prob_excellent = risk_metrics['Probability_Excellent']
                color = "#28a745" if prob_excellent > 25 else "#ffc107" if prob_excellent > 15 else "#dc3545"
                st.markdown(f"""
                <div style="background: linear-gradient(90deg, {color}20 0%, {color}10 100%);
                            padding: 1rem; border-radius: 10px; border-left: 5px solid {color};">
                    <h4 style="color: {color}; margin: 0;">Probability of Excellent Return</h4>
                    <h2 style="color: {color}; margin: 0.5rem 0;">{prob_excellent:.1f}%</h2>
                    <p style="margin: 0;">Chance of top-quartile performance</p>
                </div>
                """, unsafe_allow_html=True)

            with col3:
                downside_risk = (mc_results['NPV'] < 0).mean() * 100
                color = "#dc3545" if downside_risk > 30 else "#ffc107" if downside_risk > 15 else "#28a745"
                st.markdown(f"""
                <div style="background: linear-gradient(90deg, {color}20 0%, {color}10 100%);
                            padding: 1rem; border-radius: 10px; border-left: 5px solid {color};">
                    <h4 style="color: {color}; margin: 0;">Downside Risk</h4>
                    <h2 style="color: {color}; margin: 0.5rem 0;">{downside_risk:.1f}%</h2>
                    <p style="margin: 0;">Chance of negative NPV</p>
                </div>
                """, unsafe_allow_html=True)

            # Monte Carlo Charts
            st.subheader("📊 Monte Carlo Analysis Charts")

            fig_npv_dist, fig_correlation, fig_sensitivity = create_monte_carlo_charts(mc_results)

            # NPV Distribution
            st.plotly_chart(fig_npv_dist, use_container_width=True)

            col1, col2 = st.columns(2)

            with col1:
                # Correlation Matrix
                st.plotly_chart(fig_correlation, use_container_width=True)

            with col2:
                # Sensitivity Analysis
                st.plotly_chart(fig_sensitivity, use_container_width=True)

            # Percentile Analysis
            st.subheader("📈 Percentile Analysis")

            percentiles_df = pd.DataFrame({
                'Percentile': ['P10 (Pessimistic)', 'P25', 'P50 (Expected)', 'P75', 'P90 (Optimistic)'],
                'NPV ($000s)': [
                    risk_metrics['P10_NPV'],
                    np.percentile(mc_results['NPV'], 25),
                    risk_metrics['P50_NPV'],
                    np.percentile(mc_results['NPV'], 75),
                    risk_metrics['P90_NPV']
                ],
                'IRR (%)': [
                    np.percentile(mc_results['IRR'], 10),
                    np.percentile(mc_results['IRR'], 25),
                    np.percentile(mc_results['IRR'], 50),
                    np.percentile(mc_results['IRR'], 75),
                    np.percentile(mc_results['IRR'], 90)
                ]
            })

            percentiles_df['NPV ($000s)'] = percentiles_df['NPV ($000s)'].round(0)
            percentiles_df['IRR (%)'] = percentiles_df['IRR (%)'].round(1)

            st.dataframe(percentiles_df, use_container_width=True)

            # Store Monte Carlo results for export
            st.session_state['mc_results'] = mc_results
            st.session_state['risk_metrics'] = risk_metrics

    # Export functionality
    st.header("📊 Export Results")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("📊 Export to Excel"):
            params = {
                'Oil Price ($/bbl)': oil_price,
                'Gas Price ($/MCF)': gas_price,
                'Initial Production (bbl/month)': initial_production,
                'Monthly Decline Rate (%)': decline_rate,
                'Discount Rate (%)': discount_rate
            }

            # Include Monte Carlo results if available
            mc_results = st.session_state.get('mc_results', None)
            risk_metrics = st.session_state.get('risk_metrics', None)

            excel_data = export_to_excel(df, summary, params, mc_results, risk_metrics)

            st.download_button(
                label="📥 Download Excel File",
                data=excel_data,
                file_name=f"BAH_Jackson_Sands_Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col2:
        if st.button("📊 Export to CSV"):
            # Include Monte Carlo results if available
            mc_results = st.session_state.get('mc_results', None)
            risk_metrics = st.session_state.get('risk_metrics', None)

            params = {
                'Oil Price ($/bbl)': oil_price,
                'Gas Price ($/MCF)': gas_price,
                'Initial Production (bbl/month)': initial_production,
                'Monthly Decline Rate (%)': decline_rate,
                'Discount Rate (%)': discount_rate
            }

            csv_data = export_to_csv_professional(df, summary, params, mc_results, risk_metrics)

            st.download_button(
                label="📥 Download Professional CSV Report",
                data=csv_data,
                file_name=f"BAH_Jackson_Sands_Complete_Analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                mime="text/csv"
            )

    # Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #666; margin-top: 2rem;">
        <p><strong>BAH Jackson Sands Oil & Gas Investment Analysis</strong></p>
        <p>Professional Interactive Financial Model | Built with Streamlit & Python</p>
        <p>© 2025 Excel Mega-Builder | Real-time calculations with industry-standard parameters</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()